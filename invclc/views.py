from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils.http import urlencode
from .models import Invoice,DeletedInvoice,ModifiedInvoice,TrackingPayment,Invitation
from .forms import InvoiceForm
from django.http import JsonResponse,HttpResponseServerError,HttpResponse,HttpResponseBadRequest,HttpResponseRedirect
from django.views.decorators.csrf import csrf_exempt ,csrf_protect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q,F
from django.utils import timezone
from authentication.models import CustomUser,Person,Notification,RegisterMedicals,ConnectMedicals
from datetime import datetime
import csv
import json
from django.utils import timezone
from django.views.decorators.http import require_POST,require_GET
from django.core.management import call_command
from decimal import Decimal
from openpyxl import Workbook
from tablib import Dataset
from .forms import UploadFileForm
from django.core.serializers import serialize
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
from itsdangerous import URLSafeSerializer,BadSignature
import openpyxl
import logging
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType
from .utils import generate_tempno,RegisterUserTempNo
from authentication.service import SendNotification

from django.db.models import Sum
from calendar import month_name

logger = logging.getLogger(__name__)
def upload_csv(request):
    
    checked_username= None
    senderName = None
        
    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
        
        # Loop through each read notification
        for notification in read_notifications:
            # Get the username of the sender of the notification
            sender_username = notification.sender.username
            # Get the username of the receiver of the notification (current user)
            receiver_username = notification.receiver.username
            
            # Find all staff users with the same username as the sender
            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            # Get the non-staff user (current user) with the specified username
            normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
            
            # Loop through each staff user with the sender's username
            for user in staff_senders:
                sender = user.username

                try:
                    # Get the staff user with the current username
                    senderName = CustomUser.objects.get(username=sender, is_staff=True)
                    # Get the non-staff user (current user)
                    receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                    if receiverName and senderName:
                        # Store the username of the current non-staff user
                        checked_username = receiverName.username
                except Exception as user_error:
                    # Handle exceptions that occur while processing a specific user
                    messages.error(request, user_error)
    except Exception as general_error:  
        # Handle exceptions that occur while processing notifications
        messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))
            

            
        if request.method == 'POST' and request.FILES.get('file'):
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES['file']
                if file.name.endswith(('.csv')):
                    try:
                        csv_data = file.read().decode('utf-8').splitlines()
                        csv_reader = csv.DictReader(csv_data)
                        
                        current_user = request.user
                        
                        for row in csv_reader:
                            # Convert date string to datetime object
                            try:
                                invoice_date = datetime.strptime(row['invoice_date'], '%d/%m/%Y').date()
                            except ValueError:
                                invoice_date = datetime.strptime(row['invoice_date'], '%d-%m-%Y').date()
                            
                            # Convert string values to integers
                            invoice_amount = int(row['invoice_amount'])
                            payment_amount = int(row['payment_amount'])
                            invoice_numbers :str = row['invoice_number']

                            
                            # Calculate balance amount
                            balance_amount = invoice_amount - payment_amount
                            
                            # Get the current date and time
                            current_date = datetime.now().date()
                            current_time = datetime.now().time()
                            
                            # Create Invoice object for each row in the CSV file
                            if str(request.user) == checked_username :
                                check_invoice_number = Invoice.objects.filter(user = senderName ,invoice_number = invoice_numbers)
                                if check_invoice_number.exists():
                                    return JsonResponse({'error':'This Invoice Already Exists in Your Medical '},status = 400)
                                
                                invoice = Invoice.objects.create(
                                    user=senderName ,
                                    pharmacy_name=row['pharmacy_name'],
                                    invoice_number=row['invoice_number'],
                                    invoice_date=invoice_date,
                                    invoice_amount=invoice_amount,
                                    payment_amount=payment_amount,
                                    balance_amount=balance_amount,
                                    today_date=current_date,  # Use current date
                                    current_time=current_time,  # Use current time
                                    updated_by=row['updated_by']
                                )
                            else:
                                check_invoice = Invoice.objects.filter(user = current_user ,invoice_number = invoice_numbers)
                                if check_invoice.exists():
                                    return JsonResponse({'error':'This Invoice Already Exists in Your Medical '},status = 400)
                                
                                invoice = Invoice.objects.create(
                                    user=current_user,
                                    pharmacy_name=row['pharmacy_name'],
                                    invoice_number=row['invoice_number'],
                                    invoice_date=invoice_date,
                                    invoice_amount=invoice_amount,
                                    payment_amount=payment_amount,
                                    balance_amount=balance_amount,
                                    today_date=current_date,  # Use current date
                                    current_time=current_time,  # Use current time
                                    updated_by=row['updated_by']
                                )

                        invoice.save()
                        return JsonResponse({'message': 'CSV Data uploaded successfully'})
                    except Exception as e:
                        logger.exception("Error processing CSV file")
                        return JsonResponse({'error': f'Error processing CSV file: {str(e)}'}, status=500)
                else:
                    return JsonResponse({'error': 'File format not supported. Please upload a CSV file.'}, status=400)
            else:
                return JsonResponse({'error': 'Form is not valid'}, status=400)
        else:
            return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required(login_url='/')
def exports_to_csv(request):
    checked_username = None
    senderName = None

    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)

        for notification in read_notifications:
            sender_username = notification.sender.username
            receiver_username = notification.receiver.username

            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            normal_user = get_object_or_404(CustomUser, username=receiver_username, is_staff=False)

            for user in staff_senders:
                sender = user.username

                try:
                    senderName = get_object_or_404(CustomUser, username=sender, is_staff=True)
                    receiverName = get_object_or_404(CustomUser, username=normal_user.username, is_staff=False)

                    if receiverName and senderName:
                        checked_username = receiverName.username
                except Exception as user_error:
                    return HttpResponse(status=500)

        if checked_username == str(request.user):
            target_user = senderName
        else:
            target_user = request.user

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="Import_DataCSV_{request.user}.csv"'

        writer = csv.writer(response)
        writer.writerow(['invoice_number', 'invoice_amount', 'payment_amount', 'updated_by', 'updated_date', 'Paid/Pending Status'])

        for obj in Invoice.objects.filter(user=target_user):
            status = 'paid' if obj.balance_amount == 0 else 'pending'
            writer.writerow([obj.invoice_number, obj.invoice_amount, obj.payment_amount, obj.updated_by, obj.today_date, status])

        return response
    except Exception as e:
        return HttpResponse(status=500)

def exports_to_xlsx(request):   
    checked_username = None
    senderName = None
    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
        
        # Loop through each read notification
        for notification in read_notifications:
            # Get the username of the sender of the notification
            sender_username = notification.sender.username
            # Get the username of the receiver of the notification (current user)
            receiver_username = notification.receiver.username
            
            # Find all staff users with the same username as the sender
            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            # Get the non-staff user (current user) with the specified username
            normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
            
            # Loop through each staff user with the sender's username
            for user in staff_senders:
                sender = user.username

                try:
                    # Get the staff user with the current username
                    senderName = CustomUser.objects.get(username=sender, is_staff=True)
                    # Get the non-staff user (current user)
                    receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                    if receiverName and senderName:
                        # Store the username of the current non-staff user
                        checked_username = receiverName.username
                except Exception as user_error:
                    # Handle exceptions that occur while processing a specific user
                    messages.error(request, user_error)
    except Exception as general_error:  
        # Handle exceptions that occur while processing notifications
        messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))
        
    if checked_username == str(request.user):
        currentuser = request.user
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="ImportData{currentuser}.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['IN. No.', 'Total Amount', 'Last Payment Date', 'Payed', 'Paid/Pending'])
        for obj in Invoice.objects.filter(user=senderName):
            if obj.balance_amount == 0:
                status = 'paid'
            else:
                status = 'pending'
            worksheet.append([obj.invoice_number, obj.invoice_amount, obj.today_date, obj.payment_amount, status])

        workbook.save(response)
        return response

    else:
        currentuser = request.user
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="ImportDataXLSX_{currentuser}.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['IN. No.', 'Total Amount', 'Last Payment Date', 'Payed', 'Paid/Pending'])
        for obj in Invoice.objects.filter(user=request.user):
            if obj.balance_amount == 0:
                status = 'paid'
            else:
                status = 'pending'
            worksheet.append([obj.invoice_number, obj.invoice_amount, obj.today_date, obj.payment_amount, status])

        workbook.save(response)
        return response


@login_required(login_url='/')
def exports_to_json(request):
    checked_username = None
    senderName = None
    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
        
        # Loop through each read notification
        for notification in read_notifications:
            # Get the username of the sender of the notification
            sender_username = notification.sender.username
            # Get the username of the receiver of the notification (current user)
            receiver_username = notification.receiver.username
            
            # Find all staff users with the same username as the sender
            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            # Get the non-staff user (current user) with the specified username
            normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
            
            # Loop through each staff user with the sender's username
            for user in staff_senders:
                sender = user.username

                try:
                    # Get the staff user with the current username
                    senderName = CustomUser.objects.get(username=sender, is_staff=True)
                    # Get the non-staff user (current user)
                    receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                    if receiverName and senderName:
                        # Store the username of the current non-staff user
                        checked_username = receiverName.username
                except Exception as user_error:
                    # Handle exceptions that occur while processing a specific user
                    messages.error(request, user_error)
    except Exception as general_error:  
        # Handle exceptions that occur while processing notifications
        messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))
        
    if checked_username == str(request.user):
        invoices = Invoice.objects.filter(user= senderName )
        json_data = serialize('json', invoices)
        response = HttpResponse(json_data, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="invoices_{ senderName }.json"'
        
        return response
    else:
        current_user = request.user
        invoices = Invoice.objects.filter(user=current_user)
        json_data = serialize('json', invoices)
        response = HttpResponse(json_data, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="invoices_{current_user.username}.json"'
        
        return response

@login_required(login_url='/')
def index_view(request):
    current_user = request.user
    invoices = Invoice.objects.filter(user=current_user)
    payment_details = invoices.order_by('-id')
    q_details = Invoice.objects.filter(Q(user=current_user), ~Q(balance_amount=0.00), ~Q(balance_amount=F('invoice_amount'))).order_by('-id')
    search_details = Invoice.objects.filter(Q(user=current_user), ~Q(balance_amount=0.00), Q(payment_amount=0))
    payed_details = invoices.filter(balance_amount=0.00).order_by('-id')

    senderName = None
    check_admin = None
    unique_code_id = None
    checked_username = None

    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
        
        # Loop through each read notification
        for notification in read_notifications:
            # Get the username of the sender of the notification
            sender_username = notification.sender.username
            
            # Get the username of the receiver of the notification (current user)
            receiver_username = notification.receiver.username
            
            # Find all staff users with the same username as the sender
            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            
            if staff_senders.exists():
                # Get the non-staff user (current user) with the specified username
                normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
                
                # Loop through each staff user with the sender's username
                for user in staff_senders:
                    sender = user.username

                    try:
                        # Get the staff user with the current username
                        senderName = CustomUser.objects.get(username=sender, is_staff=True)
                        # Get the non-staff user (current user)
                        receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                        if receiverName and senderName:
                            # Store the username of the current non-staff user
                            checked_username = receiverName.username
                    except Exception as user_error:
                        # Handle exceptions that occur while processing a specific user
                        messages.error(request, user_error)
    except Exception as general_error:  
        # Handle exceptions that occur while processing notifications
        messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))

    
    profile , created =  Person.objects.get_or_create(user=current_user)
    profile.temporaryNo = RegisterUserTempNo(current_user)
    profile.save()
    
    unique_id = f"{request.user} - Please Update Your Profile"
    userCode = senderName if str(request.user) == checked_username else request.user
    try:
        unique_code = Person.objects.get(user=userCode)
        unique_id = unique_code.UniqueId
    except Person.DoesNotExist:
        unique_id = f"'{request.user}' Please Update Your Profile"

    # Set Deleted Invoice History

    if str(request.user) == checked_username:
        DeleteHistory = DeletedInvoice.objects.filter(user=senderName).order_by('-id')
    else:
        DeleteHistory = DeletedInvoice.objects.filter(user=request.user).order_by('-id')
        
    if not DeleteHistory.exists():
        DeleteHistory = "No Deletion Found"
        
    # Set Modified Invoice History
    if str(request.user) == checked_username:
        ModifiedHistory = ModifiedInvoice.objects.filter(user = senderName).order_by('-id')
    else:
        ModifiedHistory = ModifiedInvoice.objects.filter(user = request.user).order_by('-id')
        
    if not ModifiedHistory.exists():
        ModifiedHistory = "No Updatation Found"

    
    # Set NOtification Histor

    notification_history :object = ConnectMedicals.objects.filter(request_sender = request.user).order_by('-id')
    if not notification_history:
        notification_history :object = ConnectMedicals.objects.filter(request_receiver = request.user).order_by('-id')
    
    if not notification_history.exists():
        notification_history :str = "No Notification Found !!"
        
    Storename = None
    admin_invoices = None
    try:
        if checked_username == str(request.user): 
            Storename = Person.objects.get(user=senderName)
        else:
            Storename = Person.objects.get(user=request.user)
        if Storename and Storename.MedicalShopName:
            modifiedStore = convert_Medical(Storename.MedicalShopName)
        else:
            modifiedStore = "Not Found"
    except Person.DoesNotExist:
        modifiedStore = "Not Found"


    try:
        if checked_username == str(request.user): 
            Medicalname = Person.objects.get(user=senderName)
        else:
            Medicalname = Person.objects.get(user=current_user)
    except Person.DoesNotExist:
        Medicalname = ''

    entryDisable = None
    userProfile = None

    try:
        user_to_check = request.user
        userObject = CustomUser.objects.get(username=user_to_check.username)
        userPosition = userObject.position
        
        if userPosition in ['Member', 'Senior']:
            entryDisable = False
        else:
            try:
                userProfile = Person.objects.get(user=user_to_check)
                if (userProfile.MedicalShopName and
                    userProfile.DrugLiceneseNumber1 and
                    userProfile.DrugLiceneseNumber2):
                    entryDisable = False
                else:
                    entryDisable = True
            except Person.DoesNotExist:
                entryDisable = True
                messages.error(request, "Please complete your profile details to access the fields.")
                
    except CustomUser.DoesNotExist:
        entryDisable = True
        messages.error(request, "User does not exist.")

    
    full_paid =None
    edit_paid = None
    partially_paid = None
    debt_paid = None
    delete_history = None
    modified_history = None
    display_medicalname = None
    display_dl1 = None
    display_dl2 = None
    
    if checked_username == str(request.user):
        admin_invoices = Invoice.objects.filter(user=senderName)
        full_paid = admin_invoices.filter(balance_amount=0.00).order_by('-id')
        edit_paid = admin_invoices.filter().order_by('-id')
        partially_paid = admin_invoices.filter(~Q(balance_amount=0.00), ~Q(balance_amount=F('invoice_amount'))).order_by('-id')
        debt_paid = admin_invoices.filter(~Q(balance_amount=0.00), Q(payment_amount=0))
        
    else:
        admin_invoices = Invoice.objects.filter(user=request.user)
        full_paid = admin_invoices.filter(balance_amount=0.00).order_by('-id')
        edit_paid = admin_invoices.filter().order_by('-id')
        partially_paid = admin_invoices.filter(~Q(balance_amount=0.00), ~Q(balance_amount=F('invoice_amount'))).order_by('-id')
        debt_paid = admin_invoices.filter(~Q(balance_amount=0.00), Q(payment_amount=0))
        
    get_medicalname = Person.objects.get(user = request.user)
    if get_medicalname:
        is_medical_exists :object = get_medicalname.MedicalShopName
        is_dl1_exists :object = get_medicalname.DrugLiceneseNumber1
        is_dl2_exists :object = get_medicalname.DrugLiceneseNumber2
    
        
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            required_fields = [
                'pharmacy_name', 'invoice_number', 'invoice_date',
                'invoice_amount', 'payment_amount'
            ]
            
            missing_fields = [field for field in required_fields if not data.get(field)]
            if missing_fields:
                messages.error(request,f'Missing fields: {", ".join(missing_fields)}')
                return JsonResponse({
                    'success': False,
                    'message': f'Missing fields: {", ".join(missing_fields)}'
                }, status=400)

            # Convert numeric fields from strings to integers or floats
            try:
                invoice_amount = float(data['invoice_amount'])
                payment_amount = float(data['payment_amount'])
                
                if payment_amount > invoice_amount:
                    messages.error(request,'Payment Not Valid')
                    return JsonResponse({
                        'success':False,
                        'message':'Payment Not Valid '
                        },status=400)
                    
            except ValueError:
                messages.error(request,'Invoice amount and payment amount must be numeric')
                return JsonResponse({
                    'success': False,
                    'message': 'Invoice amount and payment amount must be numeric'
                }, status=400)

            
            # Convert invoice_date to YYYY-MM-DD format
            try:
                invoice_date = datetime.strptime(data['invoice_date'], '%d/%m/%Y').date()
            except ValueError:
                messages.error(request,'Invoice date must be in DD/MM/YYYY format')
                return JsonResponse({
                    'success': False,
                    'message': 'Invoice date must be in DD/MM/YYYY format'
                }, status=400)
           

            if checked_username == str(request.user):
                check_invoice_number :object = Invoice.objects.filter(user = senderName, invoice_number = data['invoice_number'])
                if check_invoice_number.exists():
                    messages.error(request,"This Invoice Number Already Exists in Your Medical")
                    return JsonResponse({'success':False,'message':'This Invoice Number Already Exists in Your Medical'},status = 400)
                
                invoice_data = Invoice(
                    user=senderName,
                    pharmacy_name=data['pharmacy_name'],  # Keep as string
                    invoice_number=data['invoice_number'],
                    invoice_date=invoice_date,  # Use converted date
                    invoice_amount=invoice_amount,  # Convert to float or int
                    payment_amount=payment_amount,  # Convert to float or int
                    today_date=datetime.now().date(),  # Set to current date
                    current_time=datetime.now().time(),  # Set to current time
                )
                invoice_data.updated_by=request.user

            else:
                check_invoice :object = Invoice.objects.filter(user = request.user, invoice_number = data['invoice_number'])
                if check_invoice.exists():
                    messages.error(request,"This Invoice Number Already Exists in Your Medical")
                    return JsonResponse({'success':False,'message':'This Invoice Number Already Exists in Your Medical'},status = 400)
                
                invoice_data = Invoice(
                    user=request.user,
                    pharmacy_name=data['pharmacy_name'],  # Keep as string
                    invoice_number=data['invoice_number'],
                    invoice_date=invoice_date,  # Use converted date
                    invoice_amount=invoice_amount,  # Convert to float or int
                    payment_amount=payment_amount,  # Convert to float or int
                    today_date=datetime.now().date(),  # Set to current date
                    current_time=datetime.now().time(),  # Set to current time
                )
                invoice_data.updated_by=request.user

            if entryDisable:
                messages.error(request,'Please Fill Up the data in Profile Page')
                return JsonResponse({
                    'success': False,
                    'message': 'Please Fill Up the data in Profile Page'
                }, status=400)
            else:
                invoice_data.save()
            
            user_to_save = senderName if checked_username == str(request.user) else request.user

            tracking_payment = TrackingPayment(
                user=user_to_save,
                Medical_name = data['pharmacy_name'],
                Bill_no = data['invoice_number'],
                Medical_payments = invoice_amount,
                payment_date = datetime.now().date(),
                paying_amount = payment_amount
            )
            tracking_payment.save()

            try:
                # Fetch the medical details
                is_medical = Person.objects.get(MedicalShopName__iexact=invoice_data.pharmacy_name)
                receiver_invoice_data = Invoice.objects.filter(user=request.user, pharmacy_name=invoice_data.pharmacy_name)
                current_user_medical = Person.objects.get(user = request.user)

                # Check if there is a connection between the users
                is_connected = ConnectMedicals.objects.filter(
                    request_sender=request.user,
                    request_receiver=is_medical.user,
                    is_read=True,
                    accept_status=True
                ).exists()

                if not is_connected:
                    is_connected = ConnectMedicals.objects.filter(
                        request_sender=is_medical.user,
                        request_receiver=request.user,
                        is_read=True,
                        accept_status=True
                    ).exists()

                if is_connected:
                    for selected_invoice in receiver_invoice_data:
                        # Check if the invoice already exists
                        invoice_exists = Invoice.objects.filter(
                            user=is_medical.user,
                            pharmacy_name=current_user_medical.MedicalShopName,
                            invoice_number=selected_invoice.invoice_number,
                            invoice_date=selected_invoice.invoice_date,
                            invoice_amount=selected_invoice.invoice_amount,
                            balance_amount=selected_invoice.balance_amount,
                            payment_amount=selected_invoice.payment_amount,
                            today_date=selected_invoice.today_date,
                            current_time=selected_invoice.current_time,
                            updated_by=selected_invoice.updated_by
                        ).exists()

                        if invoice_exists:
                            continue 

                        # Set collaborator and save
                        selected_invoice.collaborator_invoice = is_medical.user
                        selected_invoice.save()

                        # Create a new invoice for the connected user
                        Invoice.objects.create(
                            user=is_medical.user,
                            pharmacy_name=current_user_medical.MedicalShopName,
                            invoice_number=selected_invoice.invoice_number,
                            invoice_date=selected_invoice.invoice_date,
                            invoice_amount=selected_invoice.invoice_amount,
                            balance_amount=selected_invoice.balance_amount,
                            payment_amount=selected_invoice.payment_amount,
                            today_date=selected_invoice.today_date,
                            current_time=selected_invoice.current_time,
                            updated_by=selected_invoice.updated_by,
                            collaborator_invoice=request.user,
                        )

            except Person.DoesNotExist:
                messages.error(request, 'Medical shop not found')
                return JsonResponse({
                    'success': False,
                    'message': invoice_data.pharmacy_name,
                    'status':404,
                }, status=404)
            except Exception as e:
                messages.error(request, f'An error occurred: {str(e)}')
                return JsonResponse({
                    'success': False,
                    'message': f'An error occurred: {str(e)}'
                }, status=500)


            
            messages.success(request,'Invoice saved successfully!')
            return JsonResponse({'success': True, 'message': 'Invoice saved successfully!'})
        except json.JSONDecodeError:
            messages.error(request,'Invalid JSON data')
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
        except Exception as e:
            messages.error(request,str(e))
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    context = {
        'payment': payment_details,
        'payed_details': payed_details,
        'q': q_details,
        'entryDisable':entryDisable,
        'search': search_details,
        'uniqueid': unique_id,
        'DeleteHistory': DeleteHistory,
        'ModifiedHistory': ModifiedHistory,
        'medicalname': Medicalname,
        'MedicalStatus': modifiedStore,
        'check_user': checked_username,
        'check_admin': check_admin,
        'unique': unique_code_id,
        'current_user': str(request.user),
        'admin_invoice':admin_invoices,
        'full_paid': full_paid,
        'edit_paid': edit_paid,
        'partially_paid': partially_paid,
        'debt_paid': debt_paid,
        'coloborate_delete':delete_history,
        'colloborate_modified':modified_history,
        'is_medical_exists':is_medical_exists,
        'is_dl1_exists':is_dl1_exists,
        'is_dl2_exists':is_dl2_exists,
        'notification_history':notification_history,
    }
    return render(request, 'invclc/index.html', context)

@require_POST
def update_profile(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            print(data)

            pharmacy_name = data.get('pharmacy_name', "")
            dl1 = data.get('dl1', "")
            dl2 = data.get('dl2', "")
            medical_name = data.get('medicalName',"")
            invoice_number = data.get('invoice_number',"")
            # invoice_date = data.get('invoice_date',"")
            # invoice_amount = data.get('invoice_amount',"")
            # payment_amount = data.get('payment_amount',"")

            user_to_save = request.user

            # Check for duplicates
            duplicate_errors = []

            if Person.objects.exclude(user=request.user).filter(MedicalShopName=pharmacy_name).exists():
                duplicate_errors.append('Pharmacy name already exists')

            if Person.objects.exclude(user=request.user).filter(DrugLiceneseNumber1=dl1).exists():
                duplicate_errors.append('Drug license number 1 already exists')

            if Person.objects.exclude(user=request.user).filter(DrugLiceneseNumber2=dl2).exists():
                duplicate_errors.append('Drug license number 2 already exists')


            try:
                invoiceAmount = float(data['invoice_amount'])
                paymentAmount = float(data['payment_amount'])
                
                if paymentAmount > invoiceAmount:
                    messages.error(request,'Payment Not Valid')
                    return JsonResponse({
                        'success':False,
                        'message':'Payment Not Valid '
                        },status=400)
                    
            except ValueError:
                messages.error(request,'Invoice amount and payment amount must be numeric')
                return JsonResponse({
                    'success': False,
                    'message': 'Invoice amount and payment amount must be numeric'
                }, status=400)

            
            # Convert invoice_date to YYYY-MM-DD format
            try:
                invoiceDate = datetime.strptime(data['invoice_date'], '%d/%m/%Y').date()
            except ValueError:
                messages.error(request,'Invoice date must be in DD/MM/YYYY format')
                return JsonResponse({
                    'success': False,
                    'message': 'Invoice date must be in DD/MM/YYYY format'
                }, status=400)

            if Invoice.objects.filter(user=request.user,invoice_number =invoice_number).exists():
                duplicate_errors.append('Invoice Number Already Exist in Your Medical click Cancel to close')

            if duplicate_errors:
                return JsonResponse({
                    'success': False,
                    'message': ' '.join(duplicate_errors)
                }, status=400)

            profile = get_object_or_404(Person, user=user_to_save)
            
            if pharmacy_name:
                profile.MedicalShopName = pharmacy_name
                
            if dl1:
                profile.DrugLiceneseNumber1 = dl1
                
            if dl2:
                profile.DrugLiceneseNumber2 = dl2
                
            profile.save()

            Invoice.objects.create(
                user = request.user,
                pharmacy_name = medical_name,
                invoice_number = invoice_number,
                invoice_date = invoiceDate,
                invoice_amount = invoiceAmount,
                payment_amount = paymentAmount
            )

            messages.success(request, "Data added successfully.")
            return JsonResponse({'success': True, 'message': 'Profile updated successfully!'})

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
        except Exception as e:
            messages.error(request, "An error occurred while updating the profile.")
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)
    
def convert_Medical(shopname):
    
    if shopname and shopname is not None:
        words = shopname.split()
        if len(words) == 2:
            return ''.join(word[0] for word in words).upper()
        elif len(words) == 3:
            return ''.join(word[0] for word in words).upper()
        elif len(words) > 3:
            return ''.join(word[0] for word in words[:3]).upper()
        elif len(words) == 1:
            return words[0][0].upper()
        else:
            return "####"  

@login_required(login_url='/')
def update_view(request):
    invoices = Invoice.objects.filter(user=request.user).order_by('-id')
    return render(request, 'invclc/update.html', {'invoices': invoices})

@login_required(login_url='/')
def delete_page(request):
    invoices = Invoice.objects.filter(user=request.user).order_by('-id')
    return render(request, 'invclc/delete.html', {'invoices': invoices})

@login_required(login_url='/')
def check_view(request, id):
    current_user = request.user
    Storename = None  
    user_title = "Unknown"
    trackingPayment = None
    modifiedStore = None
    senderName = None
    checked_username = None
    
    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
        
        # Loop through each read notification
        for notification in read_notifications:
            # Get the username of the sender of the notification
            sender_username = notification.sender.username
            # Get the username of the receiver of the notification (current user)
            receiver_username = notification.receiver.username
            
            # Find all staff users with the same username as the sender
            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            # Get the non-staff user (current user) with the specified username
            normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
            
            # Loop through each staff user with the sender's username
            for user in staff_senders:
                sender = user.username

                try:
                    # Get the staff user with the current username
                    senderName = CustomUser.objects.get(username=sender, is_staff=True)
                    # Get the non-staff user (current user)
                    receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                    if receiverName and senderName:
                        # Store the username of the current non-staff user
                        checked_username = receiverName.username
                except Exception as user_error:
                    # Handle exceptions that occur while processing a specific user
                    messages.error(request, user_error)
    except Exception as general_error:  
        # Handle exceptions that occur while processing notifications
        messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))
        
    if str(request.user) == checked_username:
        try:
            Storename = Person.objects.get(user=senderName)
            modifiedStore = convert_Medical(Storename.MedicalShopName)
        except Person.DoesNotExist:
            modifiedStore = "Not Found"
                
        try:
            trackingPayment = TrackingPayment.objects.filter(user=senderName).order_by('-id')
        except TrackingPayment.DoesNotExist:
            trackingPayment = []
        
        try:
            userInvoice = Invoice.objects.get(id=id, user=senderName)
            if userInvoice.pharmacy_name:
                user_title = userInvoice.pharmacy_name.title()
        except Invoice.DoesNotExist:
            pass
        
        invoices = Invoice.objects.filter(user=senderName, id=id).order_by('-id')
        
    else:
        try:
            Storename = Person.objects.get(user=current_user)
            modifiedStore = convert_Medical(Storename.MedicalShopName)
        except Person.DoesNotExist:
            modifiedStore = "Not Found"
                
        try:
            trackingPayment = TrackingPayment.objects.filter(user=request.user).order_by('-id')
        except TrackingPayment.DoesNotExist:
            trackingPayment = []
        
        try:
            userInvoice = Invoice.objects.get(id=id, user=current_user)
            if userInvoice.pharmacy_name:
                user_title = userInvoice.pharmacy_name.title()
        except Invoice.DoesNotExist:
            pass

        invoices = Invoice.objects.filter(user=current_user, id=id).order_by('-id')
    
    context = {
        'invoices': invoices,
        'user_title': user_title,
        'tracking_invoices': trackingPayment,
        'store': modifiedStore,
    }
    return render(request, 'invclc/check.html', context)


# TODO: Import View ...
@login_required(login_url='/')
def import_view(request):
    upload_csv_file = UploadFileForm()
    checked_username = None
    senderName = None
    data = None
    admin_city = None  # Initialize admin_city variable
    admin_ph = None
    admin_uniqueid = None
    admin_person = None
    admin_email = None
    admin_street = None
    admin_pincode = None
    admin_dl1 = None
    admin_dl2 = None
    user_city =None
    user_ph =None
    unique_id = None
    completed_data = None
    previous_data = None
    user_email = None
    user_street = None
    user_pincode = None
    dl1 = None
    dl2 = None

    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
        
        # Loop through each read notification
        for notification in read_notifications:
            # Get the username of the sender of the notification
            sender_username = notification.sender.username
            # Get the username of the receiver of the notification (current user)
            receiver_username = notification.receiver.username
            
            # Find all staff users with the same username as the sender
            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            # Get the non-staff user (current user) with the specified username
            normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
            
            # Loop through each staff user with the sender's username
            for user in staff_senders:
                sender = user.username

                try:
                    # Get the staff user with the current username
                    senderName = CustomUser.objects.get(username=sender, is_staff=True)
                    # Get the non-staff user (current user)
                    receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                    if receiverName and senderName:
                        # Store the username of the current non-staff user
                        checked_username = receiverName.username
                except Exception as user_error:
                    # Handle exceptions that occur while processing a specific user
                    messages.error(request, user_error)
    except Exception as general_error:  
        # Handle exceptions that occur while processing notifications
        messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))
    
    if str(request.user) == checked_username:        
        try:
            person = Person.objects.get(user=senderName)
            unique_id = person.UniqueId
            admin_person = person.user.username
            admin_city = person.City
            admin_street = person.StreetNumber
            admin_pincode = person.Pincode
            admin_dl1 = person.DrugLiceneseNumber1
            admin_dl2 = person.DrugLiceneseNumber2

            get_admin_ph = CustomUser.objects.get(username=senderName) 
            admin_ph = get_admin_ph.phone_num
            admin_uniqueid = person.UniqueId
            admin_email = get_admin_ph.email

            user = Person.objects.get(user=request.user)
            user_city = user.City
            get_user_ph = CustomUser.objects.get(username=request.user)
            user_ph = get_user_ph.phone_num 
            unique_id = user.UniqueId
            user_email = get_user_ph.email
            user_street = user.StreetNumber
            user_pincode = user.Pincode
            dl1 = user.DrugLiceneseNumber1
            dl2 = user.DrugLiceneseNumber2
            
            data = Invoice.objects.filter(user=senderName).order_by('id')
            overall_data = Invoice.objects.filter(user=senderName).order_by('id')
        except Exception as a:
            return messages.error(request,"Something Wrong Please Check",a) 
          
    else:        
        try:
            user = Person.objects.get(user=request.user)
            user_city = user.City            
            get_user_ph = CustomUser.objects.get(username=request.user)
            user_ph = get_user_ph.phone_num 
            unique_id = user.UniqueId
            user_email = get_user_ph.email
            user_street = user.StreetNumber
            user_pincode = user.Pincode
            dl1 = user.DrugLiceneseNumber1
            dl2 = user.DrugLiceneseNumber2

            data = Invoice.objects.filter(user=request.user).order_by('id')
            overall_data = Invoice.objects.filter(user=request.user).order_by('-id')
        except CustomUser.DoesNotExist:
            return messages.error(request, "CustomUser does not exist")
        except Person.DoesNotExist:
            messages.error(request, "Please Update Your Profile and try agin to import export page")
            return redirect("profile")

    user_name = request.user
    if request.method == 'POST':
        completed = request.POST.get('completed', False)
        category = request.POST.get('category', '')
        others = request.POST.get('others', False)
        not_paid = request.POST.get('all',False)
        pharmacy_name = request.POST.get('pharmacyName',False) # Corrected variable name
        
        if str(request.user) == checked_username:
            try:
                if completed == 'true':
                    completed_data = list(Invoice.objects.filter(balance_amount=0, user=senderName).values('invoice_number', 'invoice_amount', 'updated_by', 'today_date', 'payment_amount', 'balance_amount'))
                    return JsonResponse({"completed_data": completed_data})
                elif not_paid == 'true':
                    not_paid_datas = list(Invoice.objects.filter(Q(user=senderName) &~Q(balance_amount=0.00) &(~Q(balance_amount=F('invoice_amount')) |Q(payment_amount=0))).order_by('-id').values('invoice_number', 'invoice_amount', 'updated_by', 'today_date', 'payment_amount', 'balance_amount'))
                    return JsonResponse({"not_paid_data": not_paid_datas})
                elif category:
                    users_with_category = list(CustomUser.objects.filter(store_type__iexact=category).values('username', 'phone_num', 'email', 'store_type'))
                    return JsonResponse({"category_list": users_with_category})
                elif others and len(others) > 1:
                    otherStores = list(CustomUser.objects.filter(store_type__icontains=others).values('username', 'phone_num', 'email', 'store_type'))
                    if not otherStores:
                        otherStoreType = CustomUser.objects.filter(other_value__icontains=others).exists()
                        if otherStoreType:
                            storeTypeList = list(CustomUser.objects.filter(other_value__icontains=others).values('username', 'phone_num', 'email', 'other_value'))
                            return JsonResponse({"storeTypeList": storeTypeList})
                    return JsonResponse({"otherStores": otherStores}) 
                
                elif pharmacy_name and pharmacy_name != '':
                    try:
                        particular_user_invoices = Invoice.objects.filter(user=senderName, pharmacy_name=pharmacy_name).values()
                        modified_invoices = []
                        for invoice in particular_user_invoices:
                            modified_invoice = {key.replace('_', ' ').capitalize(): value for key, value in invoice.items() if key not in ['id', 'user_id', 'current_time']}
                            modified_invoices.append(modified_invoice)
                        return JsonResponse({'invoices': modified_invoices})
                    except Invoice.DoesNotExist:
                        return JsonResponse({'error': 'No invoices found for the given pharmacy name'}, status=404)
                
                else:
                    previous_data = list(Invoice.objects.filter(user=senderName).values('invoice_number', 'invoice_amount', 'updated_by', 'today_date', 'payment_amount', 'balance_amount'))
                    return JsonResponse({"previous_data": previous_data})
                
            except Exception as e:
                return messages.error(request,"Something Wrong Try Again: ",e)
        else:
            try:
                if completed == 'true':
                    completed_data = list(Invoice.objects.filter(balance_amount=0, user=request.user).values('invoice_number', 'invoice_amount', 'updated_by', 'today_date', 'payment_amount', 'balance_amount'))
                    return JsonResponse({"completed_data": completed_data})
                elif not_paid == 'true':
                    not_paid_datas = list(Invoice.objects.filter(Q(user=request.user) &~Q(balance_amount=0.00) &(~Q(balance_amount=F('invoice_amount')) |Q(payment_amount=0))).order_by('-id').values('invoice_number', 'invoice_amount', 'updated_by', 'today_date', 'payment_amount', 'balance_amount'))
                    return JsonResponse({"not_paid_data": not_paid_datas})
                elif category:
                    users_with_category = list(CustomUser.objects.filter(store_type__iexact=category).values('username', 'phone_num', 'email', 'store_type'))
                    return JsonResponse({"category_list": users_with_category})
                elif others and len(others) > 1:
                    otherStores = list(CustomUser.objects.filter(store_type__icontains=others).values('username', 'phone_num', 'email', 'store_type'))
                    if not otherStores:
                        otherStoreType = CustomUser.objects.filter(other_value__icontains=others).exists()
                        if otherStoreType:
                            storeTypeList = list(CustomUser.objects.filter(other_value__icontains=others).values('username', 'phone_num', 'email', 'other_value'))
                            return JsonResponse({"storeTypeList": storeTypeList})
                    return JsonResponse({"otherStores": otherStores}) 
                
                elif pharmacy_name and pharmacy_name != '':
                    try:
                        particular_user_invoices = Invoice.objects.filter(user=request.user, pharmacy_name=pharmacy_name).values()
                        modified_invoices = []
                        for invoice in particular_user_invoices:
                            modified_invoice = {key.replace('_', ' ').capitalize(): value for key, value in invoice.items() if key not in ['id', 'user_id', 'current_time']}
                            modified_invoices.append(modified_invoice)
                        return JsonResponse({'invoices': modified_invoices})
                    except Invoice.DoesNotExist:
                        return JsonResponse({'error': 'No invoices found for the given pharmacy name'}, status=404)
                else:
                    previous_data = list(Invoice.objects.filter(user=request.user).values('invoice_number', 'invoice_amount', 'updated_by', 'today_date', 'payment_amount', 'balance_amount'))
                    return JsonResponse({"previous_data": previous_data})
            except Exception as e:
                return messages.error(request,"Something Wrong Try Again: ",e)

            

    # medicals = []  # Initialize an empty list to store medical shop names

    # for medical_user in overall_medicals:
    #     get_user_profile = medical_user.person
    #     medicals.append(get_user_profile.MedicalShopName)
        
    # print("medicals : ", medicals)

    context = {
        'datas': data,
        # 'medicals':medicals,
        'user_email':user_email,
        'user_street':user_street,
        'user_pincode':user_pincode,
        'dl1':dl1,
        'dl2':dl2,
        'admin_email':admin_email,
        'admin_pincode':admin_pincode,
        'admin_dl1':admin_dl1,
        'admin_dl2':admin_dl2,
        'admin_street':admin_street,
        'completed_data':completed_data,
        'check_user': checked_username,
        'request_user': str(request.user),
        'admin_city': admin_city,   
        'admin_ph': admin_ph,
        'admin_id': admin_uniqueid,
        'user_city': user_city,
        'user_ph': user_ph,
        'admin_person': admin_person,
        'user': user_name,
        'unique_id': unique_id,
        'form':upload_csv_file,
        'overall_data':overall_data,
    }
    return render(request, 'invclc/import-export.html', context)



# @login_required(login_url = '/')
# def payment_view(request,payment_id):
#     invoices = Invoice.objects.get(id=payment_id)
#     return render(request, 'invclc/payment.html', {'payment': invoices})


@login_required(login_url='/')
def staticspage_view(request):
    check_user = None
    plots_data = None
    plot_data = None
    error_message = None  # Variable to hold error messages

    try:
        current_user_invoices = Invoice.objects.filter(user=request.user)
        collaborator_requests = Notification.objects.filter(sender=request.user, is_read=True)
        
        for notification in collaborator_requests:
            collaborator_request_username = notification.sender.username
            get_admin_name = notification.receiver.username
            
            collaborator_request_sender = CustomUser.objects.filter(username=collaborator_request_username, is_staff=False)
            collaborator_admin = CustomUser.objects.get(username=get_admin_name, is_staff=True)
            
            for user in collaborator_request_sender:
                collaborator_sender_username = user.username
                try:
                    current_user = CustomUser.objects.get(username=collaborator_sender_username, is_staff=False)
                    if current_user and collaborator_admin:
                        check_user = current_user.username
                except Exception as a:
                    error_message = f"Something went wrong: {a}"
                    continue

        total_amount = current_user_invoices.aggregate(Sum('invoice_amount'))['invoice_amount__sum'] or 0
        payment_amount = current_user_invoices.aggregate(Sum('payment_amount'))['payment_amount__sum'] or 0
        balance_amount = current_user_invoices.aggregate(Sum('balance_amount'))['balance_amount__sum'] or 0

        # Note: You might want to do something with plots_data and plot_data variables here
        
    except Exception as e:
        error_message = f"Something went wrong: {e}"

    if str(request.user) == check_user:
        if request.method == 'GET' and 'from_date' in request.GET and 'to_date' in request.GET:
            try:
                from_date_str = request.GET.get('from_date')
                to_date_str = request.GET.get('to_date')
                from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
                
                invoices = Invoice.objects.filter(user=collaborator_admin, invoice_date__gte=from_date, invoice_date__lte=to_date)
                total_amount = invoices.aggregate(total=Sum('invoice_amount'))['total'] or 0
                payment_amount = invoices.aggregate(total=Sum('payment_amount'))['total'] or 0
                balance_amount = invoices.aggregate(total=Sum('balance_amount'))['total'] or 0
                
                return JsonResponse({'total_amount': total_amount, 
                                     'payment_amount': payment_amount, 
                                     'balance_amount': balance_amount, 
                                     })
            except Exception as e:
                error_message = f"Something went wrong: {e}"
                return JsonResponse({"error": error_message}, status=500)
        else:
            current_user = Invoice.objects.filter(user=collaborator_admin)
            total_amount = current_user.aggregate(total=Sum('invoice_amount'))['total'] or 0
            payment_amount = current_user.aggregate(total=Sum('payment_amount'))['total'] or 0
            balance_amount = current_user.aggregate(total=Sum('balance_amount'))['total'] or 0
    else:
        if request.method == 'GET' and 'from_date' in request.GET and 'to_date' in request.GET:
            try:
                from_date_str = request.GET.get('from_date')
                to_date_str = request.GET.get('to_date')
                from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
                
                invoices = Invoice.objects.filter(user=request.user, invoice_date__gte=from_date, invoice_date__lte=to_date)
                total_amount = invoices.aggregate(total=Sum('invoice_amount'))['total'] or 0
                payment_amount = invoices.aggregate(total=Sum('payment_amount'))['total'] or 0
                balance_amount = invoices.aggregate(total=Sum('balance_amount'))['total'] or 0
                
                return JsonResponse({'total_amount': total_amount,
                                     'payment_amount': payment_amount, 
                                     'balance_amount': balance_amount, 
                                      })
            except Exception as e:
                error_message = f"Something went wrong: {e}"
                return JsonResponse({"error": error_message}, status=500)
        else:
            current_user = Invoice.objects.filter(user=request.user)
            total_amount = current_user.aggregate(total=Sum('invoice_amount'))['total'] or 0
            payment_amount = current_user.aggregate(total=Sum('payment_amount'))['total'] or 0
            balance_amount = current_user.aggregate(total=Sum('balance_amount'))['total'] or 0

    context = {
        'total_amount': total_amount,
        'payment_amount': payment_amount,
        'balance_amount': balance_amount,
        'plots_data': plots_data,
        'error_message': error_message,  # Include error message in the context
    }
    return render(request, 'invclc/static.html', context)

@login_required(login_url='/')
def get_yearly_data(request):
    if request.method == 'GET' and 'year' in request.GET:
        selected_year = request.GET.get('year')
        try:
            # Filter invoices for the selected year
            invoices = Invoice.objects.filter(user=request.user, invoice_date__year=selected_year)
            
            # Calculate total amount, total paid amount, and balance amount for the year
            total_amount = invoices.aggregate(total=Sum('invoice_amount'))['total'] or 0
            payment_amount = invoices.aggregate(total=Sum('payment_amount'))['total'] or 0
            balance_amount = invoices.aggregate(total=Sum('balance_amount'))['total'] or 0
            
            # Aggregate monthly data for the selected year
            monthly_data = invoices.values('invoice_date__month').annotate(
                total_amount=Sum('invoice_amount'),
                total_payment=Sum('payment_amount'),
                total_balance=Sum('balance_amount')
            )
            
            # Convert queryset to a list of dictionaries
            monthly_data_list = [{'month': entry['invoice_date__month'], 
                                  'total_amount': entry['total_amount'],
                                  'total_payment': entry['total_payment'],
                                  'total_balance': entry['total_balance']} 
                                 for entry in monthly_data]

            # Construct the response data
            response_data = {
                'total_amount': total_amount,
                'payment_amount': payment_amount,
                'balance_amount': balance_amount,
                'monthlyData': monthly_data_list
            }
            
            return JsonResponse(response_data)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    else:
        return JsonResponse({"error": "Invalid request"}, status=400)
    

@login_required(login_url='/')
def get_monthly_data(request):
    if request.method == 'GET' and 'month' in request.GET:
        selected_month = request.GET.get('month')
        selected_year = request.GET.get('year')
        try:
            month_number = list(month_name).index(selected_month)
            if selected_year:
                invoices = Invoice.objects.filter(user=request.user, invoice_date__month=month_number, invoice_date__year=selected_year)
            else:
                invoices = Invoice.objects.filter(user=request.user, invoice_date__month=month_number)
            
            total_amount = invoices.aggregate(total=Sum('invoice_amount'))['total'] or 0
            payment_amount = invoices.aggregate(total=Sum('payment_amount'))['total'] or 0
            balance_amount = invoices.aggregate(total=Sum('balance_amount'))['total'] or 0
            
            invoice_data = invoices.values('invoice_date', 'invoice_amount', 'payment_amount', 'balance_amount')

            return JsonResponse({
                'total_amount': total_amount,
                'payment_amount': payment_amount,
                'balance_amount': balance_amount,
                'invoice_data': list(invoice_data)
            })
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    else:
        return JsonResponse({"error": "Invalid request"}, status=400)
    

@login_required(login_url='/')
def get_total_sum(request):
    try:
        year = request.GET.get('year')
        month = request.GET.get('month')
        
        if not year or not month:
            return JsonResponse({'error': 'Year and month are required.'}, status=400)
        
        month = int(month)
        year = int(year)
        
        # Assuming invoice_date is a DateTimeField
        invoices = Invoice.objects.filter(
            user=request.user,
            invoice_date__year=year,
            invoice_date__month=month
        )
        
        total_amount = invoices.aggregate(Sum('invoice_amount'))['invoice_amount__sum'] or 0
        payment_amount = invoices.aggregate(Sum('payment_amount'))['payment_amount__sum'] or 0
        balance_amount = invoices.aggregate(Sum('balance_amount'))['balance_amount__sum'] or 0
        
        invoice_data = invoices.values('invoice_date', 'invoice_amount', 'payment_amount', 'balance_amount')
        
        response_data = {
            'total_amount': total_amount,
            'payment_amount': payment_amount,
            'balance_amount': balance_amount,
            'invoice_data': list(invoice_data)
        }
        
        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    


@require_GET
def get_data(request):
    from_date_str = request.GET.get('fromDate')
    to_date_str = request.GET.get('toDate')

    try:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    try:


        # Fetch total amounts for the entire date range
        invoices = Invoice.objects.filter(invoice_date__range=[from_date, to_date])
        total_amount = invoices.aggregate(total_amount=Sum('invoice_amount'))['total_amount'] or 0
        payment_amount = invoices.aggregate(payment_amount=Sum('payment_amount'))['payment_amount'] or 0
        balance_amount = invoices.aggregate(balance_amount=Sum('balance_amount'))['balance_amount'] or 0

        monthly_data = []
        for month in range(from_date.month, to_date.month + 1):
            invoices_in_month = Invoice.objects.filter(invoice_date__month=month, invoice_date__year=from_date.year)
            total_amount_monthly = invoices_in_month.aggregate(total_amount=Sum('invoice_amount'))['total_amount'] or 0
            payment_amount_monthly = invoices_in_month.aggregate(total_amount=Sum('payment_amount'))['total_amount'] or 0
            balance_amount_monthly = invoices_in_month.aggregate(total_amount=Sum('balance_amount'))['total_amount'] or 0
            monthly_data.append({'month': month, 'totalAmount': total_amount_monthly,'paymentAmount': payment_amount_monthly, 'balanceAmount': balance_amount_monthly})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({ 
        'total_amount' : total_amount,
        'payment_amount' : payment_amount,
        'balance_amount' : balance_amount,
        'monthlyData': monthly_data
        })


@login_required(login_url='/')
def checkmore_view(request):
    current_user = request.user
    Storename = None
    modifiedStore = None
    trackingPayment = None

    senderName = None
    checked_username = None  # Define check_user here
        
    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
        
        # Loop through each read notification
        for notification in read_notifications:
            # Get the username of the sender of the notification
            sender_username = notification.sender.username
            # Get the username of the receiver of the notification (current user)
            receiver_username = notification.receiver.username
            
            # Find all staff users with the same username as the sender
            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            # Get the non-staff user (current user) with the specified username
            normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
            
            # Loop through each staff user with the sender's username
            for user in staff_senders:
                sender = user.username

                try:
                    # Get the staff user with the current username
                    senderName = CustomUser.objects.get(username=sender, is_staff=True)
                    # Get the non-staff user (current user)
                    receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                    if receiverName and senderName:
                        # Store the username of the current non-staff user
                        checked_username = receiverName.username
                except Exception as user_error:
                    # Handle exceptions that occur while processing a specific user
                    messages.error(request, user_error)
    except Exception as general_error:  
        # Handle exceptions that occur while processing notifications
        messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))
    
    if checked_username == str(request.user):
        try:
            Storename = Person.objects.get(user=senderName)
            modifiedStore = convert_Medical(Storename.MedicalShopName)
        except Person.DoesNotExist:
            modifiedStore = "Not Found"
                
        trackingPayment = TrackingPayment.objects.filter(user=senderName).order_by('-id')
        invoices = Invoice.objects.filter(user=senderName, balance_amount=0.00).order_by('-id')
    else:
        try:
            Storename = Person.objects.get(user=current_user)
            modifiedStore = convert_Medical(Storename.MedicalShopName)
        except Person.DoesNotExist:
            modifiedStore = "Not Found"
                
        trackingPayment = TrackingPayment.objects.filter(user=request.user).order_by('-id')
        invoices = Invoice.objects.filter(user=current_user, balance_amount=0.00).order_by('-id')
    
    
    context={
        'invoices': invoices,
        'MedicalStatus':modifiedStore,
        'tracking_invoices':trackingPayment,
    }
    return render(request, 'invclc/checkmore.html',context)


@login_required(login_url='/')
def paymore_view(request):
    current_user = request.user
    senderName = None
    checked_username = None  # Define check_user here
        
    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
        
        # Loop through each read notification
        for notification in read_notifications:
            # Get the username of the sender of the notification
            sender_username = notification.sender.username
            # Get the username of the receiver of the notification (current user)
            receiver_username = notification.receiver.username
            
            # Find all staff users with the same username as the sender
            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            # Get the non-staff user (current user) with the specified username
            normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
            
            # Loop through each staff user with the sender's username
            for user in staff_senders:
                sender = user.username

                try:
                    # Get the staff user with the current username
                    senderName = CustomUser.objects.get(username=sender, is_staff=True)
                    # Get the non-staff user (current user)
                    receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                    if receiverName and senderName:
                        # Store the username of the current non-staff user
                        checked_username = receiverName.username
                except Exception as user_error:
                    # Handle exceptions that occur while processing a specific user
                    messages.error(request, user_error)
    except Exception as general_error:  
        # Handle exceptions that occur while processing notifications
        messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))
    
    if checked_username == str(request.user):
        invoices = Invoice.objects.filter(Q(user=senderName ), ~Q(balance_amount=0.00), ~Q(balance_amount=F('invoice_amount'))).order_by('-id')
        trackingPayment = TrackingPayment.objects.filter(user=senderName).order_by('-id')
    else:
        trackingPayment = TrackingPayment.objects.filter(user=request.user).order_by('-id')
        invoices = Invoice.objects.filter(Q(user=current_user), ~Q(balance_amount=0.00), ~Q(balance_amount=F('invoice_amount'))).order_by('-id')
    return render(request, 'invclc/paymore.html',{'invoices': invoices,'tracking_invoices':trackingPayment})

@login_required(login_url='/')
def updatemore_view(request):
    current_user = request.user
    senderName = None
    checked_username = None  # Define check_user here
        
    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
        
        # Loop through each read notification
        for notification in read_notifications:
            # Get the username of the sender of the notification
            sender_username = notification.sender.username
            # Get the username of the receiver of the notification (current user)
            receiver_username = notification.receiver.username
            
            # Find all staff users with the same username as the sender
            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            # Get the non-staff user (current user) with the specified username
            normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
            
            # Loop through each staff user with the sender's username
            for user in staff_senders:
                sender = user.username

                try:
                    # Get the staff user with the current username
                    senderName = CustomUser.objects.get(username=sender, is_staff=True)
                    # Get the non-staff user (current user)
                    receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                    if receiverName and senderName:
                        # Store the username of the current non-staff user
                        checked_username = receiverName.username
                except Exception as user_error:
                    # Handle exceptions that occur while processing a specific user
                    messages.error(request, user_error)
    except Exception as general_error:  
        # Handle exceptions that occur while processing notifications
        messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))

    username = senderName if checked_username == str(request.user) else current_user

    invoices = Invoice.objects.filter(user = username).order_by('-id')

    return render(request, 'invclc/updatemore.html',{'invoices': invoices})

@login_required(login_url='/')
def unpaid_debt(request):
    current_user = request.user
    senderName = None
    checked_username = None  # Define check_user here
        
    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
        
        # Loop through each read notification
        for notification in read_notifications:
            # Get the username of the sender of the notification
            sender_username = notification.sender.username
            # Get the username of the receiver of the notification (current user)
            receiver_username = notification.receiver.username
            
            # Find all staff users with the same username as the sender
            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            # Get the non-staff user (current user) with the specified username
            normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
            
            # Loop through each staff user with the sender's username
            for user in staff_senders:
                sender = user.username

                try:
                    # Get the staff user with the current username
                    senderName = CustomUser.objects.get(username=sender, is_staff=True)
                    # Get the non-staff user (current user)
                    receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                    if receiverName and senderName:
                        # Store the username of the current non-staff user
                        checked_username = receiverName.username
                except Exception as user_error:
                    # Handle exceptions that occur while processing a specific user
                    messages.error(request, user_error)
    except Exception as general_error:  
        # Handle exceptions that occur while processing notifications
        messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))

    username = senderName if checked_username == str(request.user) else request.user

    invoices = Invoice.objects.filter(Q(user=username), ~Q(balance_amount=0.00), Q(payment_amount=0))
    
    return render(request, 'invclc/unpaid_debt.html',{'invoices': invoices})

def update_invoice(request, invoice_id):
    try:
        data = json.loads(request.body)
        invoice = get_object_or_404(Invoice, pk=invoice_id)
        
        # Update the fields that don't require special handling
        invoice.pharmacy_name = data.get('pharmacy_name', invoice.pharmacy_name)
        invoice.invoice_number = data.get('invoice_number',invoice.invoice_number)
        invoice.invoice_amount = Decimal(data.get('invoice_amount', invoice.invoice_amount))
        invoice.invoice_date = data.get('invoice_date', invoice.invoice_date)
        invoice.balance_amount = Decimal(data.get('balance_amount', invoice.balance_amount))

        # Perform the necessary conversions
        invoice_date = data.get('invoice_date', invoice.invoice_date)
        invoice_date = parse_date(invoice_date)

        # Update the invoice_date field
        invoice.invoice_date = invoice_date

        # Update the payment_amount based on the balance_amount
        invoice.balance_amount = invoice.invoice_amount - invoice.payment_amount

        # Save the updated Invoice
        invoice.save()
        
        checked_username = None  # Define check_user here
        senderName = None
        
        try:
            # Get all notifications sent to the current user that have been read
            read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)
            
            # Loop through each read notification
            for notification in read_notifications:
                # Get the username of the sender of the notification
                sender_username = notification.sender.username
                # Get the username of the receiver of the notification (current user)
                receiver_username = notification.receiver.username
                
                # Find all staff users with the same username as the sender
                staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
                # Get the non-staff user (current user) with the specified username
                normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)
                
                # Loop through each staff user with the sender's username
                for user in staff_senders:
                    sender = user.username

                    try:
                        # Get the staff user with the current username
                        senderName = CustomUser.objects.get(username=sender, is_staff=True)
                        # Get the non-staff user (current user)
                        receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                        if receiverName and senderName:
                            # Store the username of the current non-staff user
                            checked_username = receiverName.username
                    except Exception as user_error:
                        # Handle exceptions that occur while processing a specific user
                        messages.error(request, user_error)
        except Exception as general_error:  
            # Handle exceptions that occur while processing notifications
            messages.error(request, "Something went wrong while exporting JSON: " + str(general_error))
            
        if checked_username == str(request.user):
            modified_invoice = ModifiedInvoice(
                user=senderName,
                modified_pharmacy=invoice.pharmacy_name,
                modified_Invoice_number=f"{invoice.invoice_number}",
                modified_Invoice_date=invoice.invoice_date,
                modified_Total_amount=invoice.invoice_amount,
                modified_balance=invoice.balance_amount,
                modified_payment=invoice.payment_amount,
                modified_today_date=invoice.today_date
            )
        else:
            modified_invoice = ModifiedInvoice(
                user=request.user,
                modified_pharmacy=invoice.pharmacy_name,
                modified_Invoice_number=f"{invoice.invoice_number}",
                modified_Invoice_date=invoice.invoice_date,
                modified_Total_amount=invoice.invoice_amount,
                modified_balance=invoice.balance_amount,
                modified_payment=invoice.payment_amount,
                modified_today_date=invoice.today_date
            )

        # Save the ModifiedInvoice object
        modified_invoice.save()

        # Display success message
        messages.success(request, "Invoice successfully modified")

        return JsonResponse({'status': 'success', 'message': 'Invoice updated successfully'})
    except json.JSONDecodeError:
        # Handle JSON decoding error
        messages.error(request, "Failed to modify invoice 1: Invalid JSON data")
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'})
    except Exception as e:
        # Handle other exceptions
        messages.error(request, f"Failed to modify invoice 2: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)})


def parse_date(date_string):
    date_string = date_string.strip()
    date_formats = ['%b. %d, %Y', '%d/%m/%Y', '%d-%m-%y', '%B %d, %Y']

    for date_format in date_formats:
        try:
            return datetime.strptime(date_string, date_format).date()
        except ValueError:
            pass

    raise ValueError("Date string does not match any expected format")


    
@login_required(login_url='/')
def delete_invoice(request, invoice_id):
    try:
        # Check if the invoice exists
        invoice = Invoice.objects.get(pk=invoice_id,user=request.user)
        checked_username = None
        senderName = None

        try:
            # Get all notifications sent to the current user that have been read
            read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)

            # Loop through each read notification
            for notification in read_notifications:
                sender_username = notification.sender.username
                receiver_username = notification.receiver.username

                # Find all staff users with the same username as the sender
                staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
                normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)

                for user in staff_senders:
                    sender = user.username

                    try:
                        senderName = CustomUser.objects.get(username=sender, is_staff=True)
                        receiverName = CustomUser.objects.get(username=normal_user.username, is_staff=False)

                        if receiverName and senderName:
                            checked_username = receiverName.username
                    except CustomUser.DoesNotExist:
                        messages.error(request, "User not found.")
                    except Exception as user_error:
                        messages.error(request, f"Error processing user: {str(user_error)}")

        except Exception as general_error:  
            messages.error(request, f"Something went wrong while processing notifications: {str(general_error)}")

        # Create a DeletedInvoice object with a unique number (using timestamp)
        if checked_username == str(request.user):
            deleted_invoice = DeletedInvoice(
                user=senderName,
                pharmacy=invoice.pharmacy_name,
                number=invoice.invoice_number,
                date=invoice.invoice_date,
                amount=invoice.invoice_amount,
                balance=invoice.balance_amount,
                payment=invoice.payment_amount,
                today_date=invoice.today_date
            )
        else:
            deleted_invoice = DeletedInvoice(
                user=request.user,
                pharmacy=invoice.pharmacy_name,
                number=invoice.invoice_number,
                date=invoice.invoice_date,
                amount=invoice.invoice_amount,
                balance=invoice.balance_amount,
                payment=invoice.payment_amount,
                today_date=invoice.today_date
            )

        # Save the DeletedInvoice object
        deleted_invoice.save()

        # Delete the Invoice object
        invoice.delete()

        messages.success(request, "Deleted Success")
        return JsonResponse({'message': 'Invoice deleted successfully'},status=200)
    
    except Invoice.DoesNotExist:
        messages.error(request, "Deletion Failed: Invoice not found")
        return JsonResponse({'error': 'Invoice not found'}, status=404)

    except Exception as e:      
        messages.error(request, f"Error deleting invoice: {str(e)}")
        return JsonResponse({'error': f"Error deleting invoice: {str(e)}"}, status=500)

    
@require_POST
def pay_invoice(request, invoice_id):
    get_sender_invoice = None

    # TODO: For more Details Check Logic.txt File
    try:
        invoice = get_object_or_404(Invoice, id=invoice_id)
        

        data = json.loads(request.body)
        updated_payment_amount = Decimal(data.get('payment_amount', invoice.payment_amount))

        if updated_payment_amount <= 0:
            # Ensure payment amount is positive
            messages.error(request,"Payment amount must be greater than zero")
            return JsonResponse({'error': 'Payment amount must be greater than zero'}, status=422)

        if invoice.payment_amount + updated_payment_amount > invoice.invoice_amount:
            # Check if the total payment exceeds the invoice amount
            messages.error(request,"payment Amount Not Valid..")
            return JsonResponse({'error': 'Total payment exceeds invoice amount'}, status=409)

        invoice.payment_amount += updated_payment_amount
        if invoice.payment_amount >= invoice.invoice_amount:
            invoice.payment_amount = invoice.invoice_amount
            invoice.balance_amount = 0
        else:
            invoice.balance_amount -= updated_payment_amount

        invoice.save()
        
        tracking_payment = TrackingPayment(
            user=invoice.user,
            Medical_name = invoice.pharmacy_name,
            Bill_no = invoice.invoice_number,
            Medical_payments = invoice.payment_amount,
            payment_date = invoice.today_date,
            paying_amount = updated_payment_amount
        )
        
        tracking_payment.save()
        
        try:
            # Fetch the collaborator's invoice
            collaborator_invoice = Invoice.objects.get(id=invoice_id, user=request.user)

            # Fetch the sender's invoice
            sender_invoice = Invoice.objects.get(user=request.user, invoice_number=collaborator_invoice.invoice_number)

            # Update the collaborator's invoice if it exists
            if collaborator_invoice.collaborator_invoice:
                sender_collaborator_invoice = Invoice.objects.get(user=collaborator_invoice.collaborator_invoice, invoice_number=collaborator_invoice.invoice_number)
                sender_collaborator_invoice.payment_amount = collaborator_invoice.payment_amount
                sender_collaborator_invoice.balance_amount = collaborator_invoice.balance_amount
                sender_collaborator_invoice.save()

            # Update the sender's invoice if the current user is the sender 
            elif sender_invoice.user == request.user:
                sender_invoice.payment_amount = collaborator_invoice.payment_amount
                sender_invoice.balance_amount = collaborator_invoice.balance_amount
                sender_invoice.save()

        except Invoice.DoesNotExist:
            print("Error: Invoice not found.")
        except Exception as e:
            print(f"Error in processing invoice: {e}")

        messages.success(request,"payment Success")

        # Check the action type (Pay or Save)
        action_type = data.get('action_type', 'Pay')

        if action_type == 'Save':
            return JsonResponse({'message': 'Invoice saved successfully'})
        else:
            return JsonResponse({'message': 'Invoice updated successfully'})

    except Exception as e:
        messages.error(request,"Payment Falied")
        print("Error : ",e)
        return JsonResponse({'error': 'Internal Server Error'}, status=500)

@require_POST
def payment_invoice(request,payment_id):
    # TODO: For more Details Check Logic.txt File
    try:
        invoice = get_object_or_404(Invoice, id=payment_id)

        data = json.loads(request.body)

        pay_amount = Decimal(data.get('payment_amount', invoice.payment_amount))
        
        if pay_amount <= 0:
            # Ensure payment amount is positive
            messages.error(request,"Payment amount must be greater than zero")
            return JsonResponse({'error': 'Payment amount must be greater than zero'}, status=422)

        if invoice.payment_amount + pay_amount > invoice.invoice_amount:
            # Check if the total payment exceeds the invoice amount
            messages.error(request,"payment Amount Not Valid..")
            return JsonResponse({'error': 'Total payment exceeds invoice amount'}, status=409)
        
        invoice.payment_amount = pay_amount + invoice.payment_amount

        if invoice.payment_amount >= invoice.invoice_amount:
            invoice.payment_amount = invoice.invoice_amount

        if invoice.balance_amount <= 0:
            invoice.balance_amount = 0

        invoice.save()
        
        tracking_payment = TrackingPayment(
            user=invoice.user,
            Medical_name = invoice.pharmacy_name,
            Bill_no = invoice.invoice_number,
            Medical_payments = invoice.payment_amount,
            payment_date = invoice.today_date,
            paying_amount = pay_amount
        )
        
        tracking_payment.save()


        try:
            # Fetch the collaborator's invoice
            collaborator_invoice = Invoice.objects.get(id=payment_id, user=request.user)

            # Fetch the sender's invoice
            sender_invoice = Invoice.objects.get(user=request.user, invoice_number=collaborator_invoice.invoice_number)

            # Update the collaborator's invoice if it exists
            if collaborator_invoice.collaborator_invoice:
                sender_collaborator_invoice = Invoice.objects.get(user=collaborator_invoice.collaborator_invoice, invoice_number=collaborator_invoice.invoice_number)
                sender_collaborator_invoice.payment_amount = collaborator_invoice.payment_amount
                sender_collaborator_invoice.balance_amount = collaborator_invoice.balance_amount
                sender_collaborator_invoice.save()

            # Update the sender's invoice if the current user is the sender 
            elif sender_invoice.user == request.user:
                sender_invoice.payment_amount = collaborator_invoice.payment_amount
                sender_invoice.balance_amount = collaborator_invoice.balance_amount
                sender_invoice.save()


        except Invoice.DoesNotExist:
            print("Error: Invoice not found.")
        except Exception as e:
            print(f"Error in processing invoice: {e}")

        messages.success(request,"payment Success")
        
        # Check the action type (Pay or Save)
        action_type = data.get('action_type', 'Pay')

        if action_type == 'Save':
            return JsonResponse({'message': 'Invoice saved successfully'})
        else:
            return JsonResponse({'message': 'Invoice updated successfully'})

    except Exception as e:
        messages.error(request,"payment Failed",e)
        return JsonResponse({'error': 'Internal Server Error'}, status=500)


def empty_csv(request):
    # Define header row
    header = ['pharmacy_name', 'invoice_number', 'invoice_date', 'invoice_amount', 'payment_amount', 'updated_by']
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_data.csv"'

    # Write header row to CSV
    writer = csv.writer(response)
    writer.writerow(header)

    return response

def empty_xlsx(request):
    # Define header row
    header = ['pharmacy_name', 'invoice_number', 'invoice_date', 'invoice_amount', 'payment_amount', 'updated_by']
    
    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write header row to worksheet
    worksheet.append(header)

    # Create HttpResponse object with XLSX content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sample_data.xlsx"'

    # Save workbook to HttpResponse
    workbook.save(response)

    return response


@login_required(login_url='/')
def admin_access(request):
    context = {}
    table_data = []
    checked_username = None
    sender_name = None
    is_admin_user = None
    unique_keys = set()  # Set to store unique keys
    checked = None  # Initialize the checked variable
    user_type = None

    try:
        # Get all notifications sent to the current user that have been read
        read_notifications = Notification.objects.filter(receiver=request.user, is_read=True)

        for notification in read_notifications:
            sender_username = notification.sender.username
            receiver_username = notification.receiver.username

            staff_senders = CustomUser.objects.filter(username=sender_username, is_staff=True)
            normal_user = CustomUser.objects.get(username=receiver_username, is_staff=False)

            for user in staff_senders:
                sender = user.username
                try:
                    sender_name = CustomUser.objects.get(username=sender, is_staff=True)
                    receiver_name = CustomUser.objects.get(username=normal_user.username, is_staff=False)
                    if receiver_name and sender_name:
                        checked_username = receiver_name.username
                except CustomUser.DoesNotExist as user_error:
                    messages.error(request, f"User error: {user_error}")
    except Exception as general_error:
        messages.error(request, f"Something went wrong while exporting JSON: {general_error}")

    context['checked_username'] = checked_username
    context['currentUser'] = str(request.user)

    try:
        user_data = sender_name if checked_username == str(request.user) else request.user
        get_all_invoice = Invoice.objects.filter(user=user_data)
        
        try:
            current_user = get_object_or_404(CustomUser, username=user_data)
            colaborator = Notification.objects.filter(sender=user_data, is_read=True)
            AdminPosition = Notification.objects.filter(sender=user_data, receiver=request.user, is_read=True).first()

            context['user_phone'] = current_user.phone_num
            context['user_username'] = current_user.username
            context['user_position'] = current_user.position
            context['colaborator'] = colaborator

            if AdminPosition:
                context['admin_position'] = AdminPosition.sender.position

        except CustomUser.DoesNotExist as e:
            context['user_phone'] = None
            context['user_username'] = None
            context['user_position'] = None
            context['admin_position'] = None

        if not get_all_invoice.exists():
            context['error'] = "No invoices found for the user."
        
        get_current_user_position = CustomUser.objects.get(username=request.user)
        context['position'] = get_current_user_position.position

        for idx, invoice in enumerate(get_all_invoice, start=1):
            try:
                profile_data = Person.objects.get(MedicalShopName=invoice.pharmacy_name)
                current_medicalname = Person.objects.get(user=request.user)
                unique_code = profile_data.UniqueId
                user_position = profile_data.user.position

                if user_position == "Admin":
                    is_admin_user = profile_data.user.username
                else:
                    is_admin_user = None  # Ensure this variable is define

                get_user = CustomUser.objects.get(username = profile_data.user)
                user_type = get_user.store_type
                if user_type == 'others':
                    user_type = get_user.other_value

                unique_key = (
                    profile_data.MedicalShopName, 
                    profile_data.DrugLiceneseNumber1, 
                    profile_data.DrugLiceneseNumber2, 
                    is_admin_user, 
                    unique_code
                )

                is_collaborate_medical = ConnectMedicals.objects.filter(
                    request_receiver=request.user,
                    request_sender=profile_data.user,
                    is_read=True,
                    accept_status=True
                )

                if not is_collaborate_medical.exists():
                    # Check the other direction of collaboration
                    is_collaborate_medical = ConnectMedicals.objects.filter(
                        request_sender=request.user,
                        request_receiver=profile_data.user,
                        is_read=True,
                        accept_status=True
                    )

                checked = False
                if is_collaborate_medical.exists():
                    checked = True

                if current_medicalname.MedicalShopName == invoice.pharmacy_name:
                    checked = True



                if unique_key not in unique_keys:
                    unique_keys.add(unique_key)
                    table_data.append({
                        's_no': idx,
                        'name': profile_data.MedicalShopName,
                        'dl_number1': profile_data.DrugLiceneseNumber1,
                        'dl_number2': profile_data.DrugLiceneseNumber2,
                        'admin_name': is_admin_user,
                        'temp_no': None,
                        'unique_no': unique_code,
                        'uique_Faild': False,
                        'generate_link': False,
                        'status': 'Active',
                        'checked': checked,
                        'user_type':user_type,
                    })
            except Person.DoesNotExist:
                temp_no = generate_tempno(invoice.pharmacy_name, invoice.id)
                check_Medical = None
                check_dl1 = None
                check_dl2 = None

                try:
                    user_data = Person.objects.get(MedicalShopName=invoice.pharmacy_name)
                    get_user = CustomUser.objects.get(user = user_data.user)
                    user_type = get_user.store_type
                    if user_type == 'others':
                        user_type = get_user.other_value
                except Exception as e:
                    user_type = None

                try:
                    check_data_medical = RegisterMedicals.objects.filter(Medical_name=invoice.pharmacy_name)
                    if check_data_medical.exists():
                        for register_medical in check_data_medical:
                            check_Medical = register_medical.Medical_name
                            check_dl1 = register_medical.dl_number1
                            check_dl2 = register_medical.dl_number2
                    else:
                        check_Medical = invoice.pharmacy_name
                        check_dl1 = None                        
                        check_dl2 = None

                except RegisterMedicals.DoesNotExist:
                    check_Medical = invoice.pharmacy_name
                    check_dl1 = None
                    check_dl2 = None
                
                unique_key = (check_Medical, check_dl1, check_dl2)
                
                if unique_key not in unique_keys:
                    unique_keys.add(unique_key)
                    table_data.append({
                        's_no': idx,
                        'name': check_Medical,
                        'dl_number1': check_dl1,
                        'dl_number2': check_dl2,
                        'admin_name': None,
                        'temp_no': temp_no,
                        'unique_no': None,
                        'generate_link': True,
                        'status': 'Inactive',
                        'checked': None,
                        'user_type':user_type,
                    })
    except Exception as e:
        context['error'] = f"Error: {e}"

    context['table_data'] = table_data   

    if request.method == "POST":
        data = json.loads(request.body)
        receiver_name = data.get('add_name')
        receiver_email = data.get('add_email')
        receiver_phone = data.get('add_number')
        receiver_position = data.get('add_position')

        try:
            if receiver_name:
                if Invitation.objects.filter(mail_sendername=request.user, mail_receiver_name=receiver_name).exists():
                    response_data = {'message': 'You have already sent a request to this receiver', 'adminName': receiver_name}
                    return JsonResponse({'error': response_data}, status=500)
                else:
                    invitation = Invitation(
                        user=request.user,
                        mail_sendername=request.user,
                        mail_receiver_name=receiver_name,
                        mail_receiver_email=receiver_email,
                        mail_receiver_phonenumber=receiver_phone,
                        mail_receiver_position=receiver_position
                    )
                    invitation.save()

                subject = "Invitation to Join Our Platform"
                text_message = f"Dear {receiver_name},\n\nYou have been invited to join our platform as {receiver_position}. Please use the following details to access your account.\n\nBest regards,\nThe Team"

                signer = URLSafeSerializer(settings.SECRET_KEY)
                data_to_sign = {
                    'sendername': str(request.user),
                    'username': receiver_name,
                    'useremail': receiver_email,
                    'userphonenumber': receiver_phone,
                    'userposition': receiver_position
                }
                signed_data = signer.dumps(data_to_sign)
                
                base_signup_url = 'https://medregia.com/invite/'
                invite_url = f"{base_signup_url}?data={signed_data}"

                html_message = render_to_string('invitation_email.html', {
                    'user_name': receiver_name,
                    'sender_mail': settings.DEFAULT_FROM_EMAIL,
                    'sender_name': request.user,
                    'signup_url': invite_url,
                    'position': receiver_position,
                })

                send_mail(
                    subject,
                    text_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [receiver_email],
                    fail_silently=False,
                    html_message=html_message
                )
                
                return JsonResponse({'status': 'success', 'message': 'Email sent successfully', 'invite_link': invite_url})
        except Exception as e:
            response_data = {'message': str(e)}
            return JsonResponse({'error': response_data}, status=500)

    return render(request, 'invclc/admin_acess.html', context)


@require_GET
def invite_user(request):
    # Extract data from URL query parameters
    encoded_data = request.GET.get('data', '')

    try:
        # Decode the encoded data using your secret key
        signer = URLSafeSerializer(settings.SECRET_KEY)
        decoded_data = signer.loads(encoded_data)

        user_position = decoded_data.get('userposition')
        sender_name = decoded_data.get('sendername')
        username = decoded_data.get('username')
        useremail = decoded_data.get('useremail')
        userphonenumber = decoded_data.get('userphonenumber')

        # Use these variables in your template or processing logic
        context = {
            'userposition': user_position,
            'sendername': sender_name,
            'username': username,
            'useremail': useremail,
            'userphonenumber': userphonenumber,
        }

        return render(request, 'invite_user.html', context)

    except BadSignature:
        return JsonResponse({'status': 'error', 'message': 'Invalid or expired invitation link'})

@csrf_exempt
@require_POST
def process_invite(request):
    try:
        # Extract data from JSON request body
        data = json.loads(request.body)
        
        new_username = data.get('new_username')
        new_useremail = data.get('new_useremail')
        new_userphonenumber = data.get('new_userphonenumber')
        new_userpassword = data.get('new_userpassword')
        new_userconfirmpassword = data.get('new_userconfirmpassword')
        new_userpin_str = data.get('new_userpin')
        new_usertype = data.get('new_usertype')
        new_userothertype = data.get('new_userothertype')
        new_userposition = data.get('new_userposition')
        new_sendername = data.get('new_sendername')

        # Validate PIN
        try:
            new_userpin = int(new_userpin_str)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': "Field 'new_userpin' expected a number but got '{}'.".format(new_userpin_str)})
        
        # Validate password match
        if new_userpassword != new_userconfirmpassword:
            return JsonResponse({'status': 'error', 'field': 'new_userconfirmpassword', 'message': 'Passwords do not match'})

        # Check if username or email already exists
        if CustomUser.objects.filter(username=new_username).exists():
            return JsonResponse({'status': 'error', 'field': 'new_username', 'message': 'Username already exists'})

        if CustomUser.objects.filter(email=new_useremail).exists():
            return JsonResponse({'status': 'error', 'field': 'new_useremail', 'message': 'Email already exists'})

        # Fetch the sender user object
        try:
            sender_user = CustomUser.objects.get(username=new_sendername)
        except CustomUser.DoesNotExist:
            return JsonResponse({'status': 'error', 'field': 'new_sendername', 'message': 'Sender does not exist'})

        # Hash the password
        hashed_password = make_password(new_userconfirmpassword)

        # Create the new CustomUser object
        new_user = CustomUser.objects.create(
            username=new_username,
            password=hashed_password,
            email=new_useremail,
            phone_num=new_userphonenumber,
            pin=new_userpin,
            store_type=new_usertype,
            other_value=new_userothertype,
            position=new_userposition,
            is_staff=True if new_userposition == 'Admin' else False
        )

        # Save the new user
        new_user.save()
        
        # Create notification
        notify = Notification(
            sender=sender_user,
            receiver=new_user,
            is_read=True,
            request_status=True,
            position=new_userposition,
        )
        notify.save()

        # Handle collaborator requests and group assignments
        if new_userposition in ['Member', 'Senior']:
            collaborator_requests = Notification.objects.filter(receiver=new_user, is_read=True)

            if collaborator_requests:
                for collaborator in collaborator_requests:
                    receiver = collaborator.receiver
                    sender = collaborator.sender

                    if new_username == receiver.username:
                        try:
                            admin_group = Group.objects.get(name='Admin Group')
                            receiver.groups.remove(admin_group)
                            receiver.is_staff = False
                            receiver.position = collaborator.position
                            receiver.save()

                            grand_accesses = Invoice.objects.filter(user=receiver)
                            tracking_access = TrackingPayment.objects.filter(user=receiver)

                            for grand_access in grand_accesses:
                                grand_access.user = sender
                                grand_access.save()

                            for tracking in tracking_access:
                                tracking.user = sender
                                tracking.save()

                            sender.save()
                            collaborator.is_read = True
                            collaborator.save()
                        except Exception as e:
                            messages.error(request, str(e))
                    else:
                        messages.error(request, "You are not authorized to become an admin.")
            else:
                messages.error(request, "There are no pending collaborator requests.")

        # Assign user to admin group with specific permissions
        if new_userposition == 'Admin' and new_user.is_staff:
            user_group, created = Group.objects.get_or_create(name="Admin Group")

            models_and_permissions = [
                (DeletedInvoice, ['view_deletedinvoice', 'delete_deletedinvoice']),
                (Invoice, ['add_invoice', 'view_invoice', 'change_invoice', 'delete_invoice']),
                (ModifiedInvoice, ['view_modifiedinvoice', 'delete_modifiedinvoice']),
                (TrackingPayment, ['view_trackingpayment', 'delete_trackingpayment']),
            ]

            for model, perms in models_and_permissions:
                content_type = ContentType.objects.get_for_model(model)
                for perm in perms:
                    permission = Permission.objects.get(codename=perm)
                    user_group.permissions.add(permission)

            new_user.groups.add(user_group)
            new_user.save()

        # Send welcome email
        subject = 'Welcome to MedRegia!'
        message = render_to_string('authentication/welcome_email.html', {'user': new_user})
        email_from = settings.DEFAULT_FROM_EMAIL
        recipient_list = [new_user.email]
        send_mail(subject, message, email_from, recipient_list)

        messages.success(request, f"Hey, {new_username}! Welcome to MedregiA.")
        return JsonResponse({'status': 'success', 'redirect_url': reverse('login')})

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@require_POST
def connect_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Refresh the page again.', 'data': None}, status=403)

        name = data.get('medicalName')
        dl_number1 = data.get('dl1')
        dl_number2 = data.get('dl2')
        unique_number = data.get('UniqueNo')
        
        try:
            check_currentuser_uniqueId = Person.objects.get(user=request.user)
        except Person.DoesNotExist:
            return JsonResponse({'message': 'User profile not found.', 'data': None}, status=404)

        check_pending_requests = ConnectMedicals.objects.filter(request_sender=request.user, is_read=False, accept_status=True)
        check_request = Person.objects.get(user = request.user)
        

        if name == check_request.MedicalShopName:
            return JsonResponse({'message':'Cannot Sent Request to Yourself '},status = 403)

        if not dl_number1 or not dl_number2:
            return JsonResponse({'message': 'Both DL numbers are required.', 'data': data}, status=400)
        
        if dl_number1 == 'None' and dl_number2 == 'None':
            return JsonResponse({'message': 'Please enter DL numbers.', 'data': data, 'Inputpopup': True}, status=400)
        
        elif dl_number1 == 'None':
            return JsonResponse({'message': 'Please enter DL number 1.', 'data': data, 'Inputpopup': True}, status=400)
        
        elif dl_number2 == 'None':
            return JsonResponse({'message': 'Please enter DL number 2.', 'data': data, 'Inputpopup': True}, status=400)
        
        try:
            person = Person.objects.get(MedicalShopName__iexact=name)
            
            # Check and update DL numbers if not set
            if not person.DrugLiceneseNumber1 or not person.DrugLiceneseNumber2:
                person.DrugLiceneseNumber1 = person.DrugLiceneseNumber1 or dl_number1
                person.DrugLiceneseNumber2 = person.DrugLiceneseNumber2 or dl_number2
                person.save()
                                                                                    
        except Person.DoesNotExist:
            return JsonResponse({
                'message': 'No medical shop found with the given name and DL numbers. Click OK to create a new record, or Cancel to abort.',
                'data': data,
                'status': 404,
                'Inputpopup': True
            }, status=404)

        if not check_currentuser_uniqueId.UniqueId:
            return JsonResponse({'message':'Please Complete Your Profile then Send the request'}, status=403)

        if check_pending_requests.exists():
            pending_for_user = any(request for request in check_pending_requests if request.request_receiver == person.user)
            if pending_for_user:
                return JsonResponse({'message': "Request not sent until the user accepts the previous request."}, status=403)

        user_uniqueId = person.UniqueId
        
        if not "###" in user_uniqueId:
            if user_uniqueId == unique_number:
                notification_message = 'Collaboration request sent successfully.'
                status = "Pending"
                SendNotification(request.user, person.user, notification_message,status)
                return JsonResponse({'message': notification_message, 'Inputpopup': False})
            else:
                return JsonResponse({'message': "Unique ID does not match with user's Unique ID.", 'data': data})
        else:
            update_profile_message = f"{name} doesn't have a Unique Id, but request sent."
            status_failed = "No UniqueId"
            SendNotification(request.user, person.user, update_profile_message,status_failed)
            return JsonResponse({'message': update_profile_message, 'data': data, 'Inputpopup': True}, status=403)

    return JsonResponse({'message': 'Invalid request method.', 'data': None, 'Inputpopup': False}, status=405)

@require_POST
def create_medical_record(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'message': 'Invalid JSON.'}, status=400)

        name = data.get('medicalName')
        dl_number1 = data.get('dl1')
        dl_number2 = data.get('dl2')
        unique_number = data.get('UniqueNo')

        # Create new Person record
        check_register = RegisterMedicals.objects.filter(Medical_name = name , dl_number1 = dl_number1, dl_number2 = dl_number2, UniqueId = unique_number)

        if check_register.exists():
            return JsonResponse({'message':"This Medical Already Register in our Records "},status=405)

        if RegisterMedicals.objects.exclude(user=request.user).filter(dl_number1=dl_number1).exists():
            return JsonResponse({'message':'Drug license number 1 already exists'},status=405)

        if RegisterMedicals.objects.exclude(user=request.user).filter(dl_number2=dl_number2).exists():
            return JsonResponse({'message':'Drug license number 2 already exists'},status=405)


        person = RegisterMedicals.objects.create(
            user=request.user,
            Medical_name=name,
            dl_number1=dl_number1,
            dl_number2=dl_number2,
            UniqueId=unique_number
        )
        person.save()
        
        return JsonResponse({'message': 'New medical record created successfully.'})

    return JsonResponse({'message': 'Invalid request method.'}, status=405)


@login_required
@require_POST
def medical_search(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'message': 'Invalid JSON.'}, status=400)

    medicalNameSearch = data.get('medicalNameSearch', '')

    try:
        if not medicalNameSearch:
            response_data = {"message": "No Results Found"}
        else:
            # Filter invoices based on user and search term
            results = Invoice.objects.filter(user=request.user, pharmacy_name__icontains=medicalNameSearch)

            if results.exists():
                # Extract pharmacy names from the filtered invoices
                pharmacy_names = results.values_list('pharmacy_name', flat=True)
                # Filter persons based on the pharmacy names
                get_medical_profile = Person.objects.filter(MedicalShopName__in=pharmacy_names)

                if get_medical_profile.exists():
                    response_data = {
                        "message": "Results Found",
                        "results": [
                            {
                                "medicals_name": profile.MedicalShopName,
                                "dlnumber_1": profile.DrugLiceneseNumber1,
                                "dlnumber_2": profile.DrugLiceneseNumber2,
                            }
                            for profile in get_medical_profile 
                        ]
                    }
                else:
                    response_data = {
                        "message": "Medical name found, but DL numbers not found",
                        "results": [
                            {
                                "medicals_name": name,
                                "dlnumber_1": "Not Found",
                                "dlnumber_2": "Not Found",
                            }
                            for name in pharmacy_names
                        ]
                    }
            else:
                response_data = {"message": "New"}

        return JsonResponse(response_data)

    except Exception as e:
        print(e)
        return JsonResponse({"message": "An error occurred"}, status=500)


